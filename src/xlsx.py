from collections import OrderedDict
from score import score_to_number
import openpyxl
import pandas as pd

def read_xlsx_to_empty_dataframe(book_path, sheet_name, skip_rows=0):
    book = openpyxl.load_workbook(book_path)
    sheet = book[sheet_name]

    columns = []

    for index, row in enumerate(sheet.rows):
        if index < skip_rows:
            continue

        columns.append(row[2].value)

    return pd.DataFrame(columns=columns)

def read_xlsx_to_dataframe(book_path, skill_sheet_name, skip_rows=0, score_sheet_name=None):
    book = openpyxl.load_workbook(book_path)
    skill_sheet = book[skill_sheet_name]

    data = OrderedDict()

    for index, row in enumerate(skill_sheet.rows):
        if index < skip_rows:
            continue

        data[row[2].value] = [int(row[3].value)]

    if (score_sheet_name is not None):
        score_sheet = book[score_sheet_name]
        data['score'] = score_to_number(score_sheet[2][1].value)

    return pd.DataFrame(data=data)

def write_score_to_xlsx(book_path, score_sheet_name, index, model_name, score):
    book = openpyxl.load_workbook(book_path)
    score_sheet = book[score_sheet_name]

    score_sheet[2 + index][0].value = model_name
    score_sheet[2 + index][1].value = score

    book.save(book_path)
